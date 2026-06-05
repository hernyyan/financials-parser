"""
Layer 1 template and source layout endpoints.

GET  /companies/{company_id}/layer1-templates/{statement_type}           — fetch stored template
POST /companies/{company_id}/layer1-templates/{statement_type}           — upsert template
POST /companies/{company_id}/layer1-templates/{statement_type}/check-layout  — LCS diff vs stored layout
POST /companies/{company_id}/layer1-templates/{statement_type}/save-layout   — upsert source layout
POST /companies/{company_id}/layer1-templates/{statement_type}/apply-changes — retroactive dataset updates
"""
import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import load_workbook

from app.config import COMPANY_DATASETS_DIR
from app.db.database import get_db
from app.models.schemas import Layer1TemplateResponse
from app.services.layout_diff import diff_layouts

logger = logging.getLogger(__name__)

router = APIRouter()

_VALID_TYPES = {"income_statement", "balance_sheet", "cash_flow_statement"}
_DATASET_SHEET = "Financials"


# ── Request models ────────────────────────────────────────────────────────────

class LayoutRow(BaseModel):
    row_index: int
    label: str


class CheckLayoutRequest(BaseModel):
    layout_rows: List[LayoutRow]


class SaveLayoutRequest(BaseModel):
    layout_rows: List[LayoutRow]


class RenameItem(BaseModel):
    old_label: str
    new_label: str


class ApplyChangesRequest(BaseModel):
    renames: List[RenameItem] = []
    additions: List[str] = []
    deletions: List[str] = []


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_dirname(name: str) -> str:
    return re.sub(r"[^\w\s\-]", "", name).strip().replace(" ", "_")


def _get_company_name(company_id: int, db: Session) -> Optional[str]:
    row = db.execute(
        text("SELECT name FROM companies WHERE id = :cid"),
        {"cid": company_id},
    ).fetchone()
    return row[0] if row else None


def _load_stored_layout(company_id: int, statement_type: str, db: Session) -> Optional[List[Dict]]:
    row = db.execute(
        text(
            "SELECT layout FROM source_layouts "
            "WHERE company_id = :cid AND statement_type = :st"
        ),
        {"cid": company_id, "st": statement_type},
    ).fetchone()
    if not row:
        return None
    layout = row[0]
    if isinstance(layout, str):
        layout = json.loads(layout)
    return layout


def _upsert_layout(company_id: int, statement_type: str, layout_rows: List[Dict], db: Session) -> None:
    layout_json = json.dumps(layout_rows)
    result = db.execute(
        text(
            "UPDATE source_layouts SET layout = :layout, updated_at = CURRENT_TIMESTAMP "
            "WHERE company_id = :cid AND statement_type = :st"
        ),
        {"layout": layout_json, "cid": company_id, "st": statement_type},
    )
    if result.rowcount == 0:
        db.execute(
            text(
                "INSERT INTO source_layouts (company_id, statement_type, layout) "
                "VALUES (:cid, :st, :layout)"
            ),
            {"cid": company_id, "st": statement_type, "layout": layout_json},
        )
    db.commit()


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/companies/{company_id}/layer1-templates/{statement_type}", response_model=Layer1TemplateResponse)
def get_layer1_template(
    company_id: int,
    statement_type: str,
    db: Session = Depends(get_db),
):
    if statement_type not in _VALID_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid statement_type: {statement_type}")

    row = db.execute(
        text(
            "SELECT id, company_id, statement_type, template, created_at, updated_at "
            "FROM layer1_templates WHERE company_id = :cid AND statement_type = :st"
        ),
        {"cid": company_id, "st": statement_type},
    ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="No template found for this company and statement type.")

    template = row[3]
    if isinstance(template, str):
        template = json.loads(template)

    return Layer1TemplateResponse(
        id=row[0],
        company_id=row[1],
        statement_type=row[2],
        template=template,
        created_at=str(row[4]) if row[4] else None,
        updated_at=str(row[5]) if row[5] else None,
    )


@router.post("/companies/{company_id}/layer1-templates/{statement_type}")
def upsert_layer1_template(
    company_id: int,
    statement_type: str,
    payload: dict,
    db: Session = Depends(get_db),
):
    if statement_type not in _VALID_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid statement_type: {statement_type}")

    template_json = json.dumps(payload)

    result = db.execute(
        text(
            "UPDATE layer1_templates SET template = :tmpl, updated_at = CURRENT_TIMESTAMP "
            "WHERE company_id = :cid AND statement_type = :st"
        ),
        {"tmpl": template_json, "cid": company_id, "st": statement_type},
    )
    if result.rowcount == 0:
        db.execute(
            text(
                "INSERT INTO layer1_templates (company_id, statement_type, template) "
                "VALUES (:cid, :st, :tmpl)"
            ),
            {"cid": company_id, "st": statement_type, "tmpl": template_json},
        )

    db.commit()
    return {"success": True}


@router.post("/companies/{company_id}/layer1-templates/{statement_type}/check-layout")
def check_layout(
    company_id: int,
    statement_type: str,
    request: CheckLayoutRequest,
    db: Session = Depends(get_db),
):
    """
    Compare incoming layout rows against the stored source layout using LCS diff.
    Returns diff result so the frontend can decide whether to show reconciliation UI.
    """
    if statement_type not in _VALID_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid statement_type: {statement_type}")

    # Check whether a template exists at all
    template_row = db.execute(
        text("SELECT id FROM layer1_templates WHERE company_id = :cid AND statement_type = :st"),
        {"cid": company_id, "st": statement_type},
    ).fetchone()
    has_template = template_row is not None

    stored_layout = _load_stored_layout(company_id, statement_type, db)

    if stored_layout is None:
        return {
            "has_template": has_template,
            "has_layout": False,
            "has_real_diff": False,
            "silent_update": False,
            "changes": [],
        }

    incoming = [{"row_index": r.row_index, "label": r.label} for r in request.layout_rows]
    diff = diff_layouts(stored_layout, incoming)

    return {
        "has_template": has_template,
        "has_layout": True,
        **diff,
    }


@router.post("/companies/{company_id}/layer1-templates/{statement_type}/save-layout")
def save_layout(
    company_id: int,
    statement_type: str,
    request: SaveLayoutRequest,
    db: Session = Depends(get_db),
):
    """Upsert the source layout record for a company/statement."""
    if statement_type not in _VALID_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid statement_type: {statement_type}")

    layout_rows = [{"row_index": r.row_index, "label": r.label} for r in request.layout_rows]
    _upsert_layout(company_id, statement_type, layout_rows, db)
    return {"success": True}


@router.post("/companies/{company_id}/layer1-templates/{statement_type}/apply-changes")
def apply_template_changes(
    company_id: int,
    statement_type: str,
    request: ApplyChangesRequest,
    db: Session = Depends(get_db),
):
    """
    Retroactively update company dataset Excel files when the template changes:
    - Renames: update label text in dataset rows
    - Additions: insert new row with null values for all existing periods
    - Deletions: remove rows from dataset

    Returns count of rows affected across all year files.
    """
    if statement_type not in _VALID_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid statement_type: {statement_type}")

    if not request.renames and not request.additions and not request.deletions:
        return {"success": True, "rows_updated": 0}

    company_name = _get_company_name(company_id, db)
    if not company_name:
        raise HTTPException(status_code=404, detail="Company not found.")

    company_dir = COMPANY_DATASETS_DIR / _safe_dirname(company_name)
    if not company_dir.exists():
        # No dataset yet — nothing to update
        return {"success": True, "rows_updated": 0}

    xlsx_files = sorted(company_dir.glob("*.xlsx"))
    total_updated = 0

    for xlsx_path in xlsx_files:
        try:
            wb = load_workbook(str(xlsx_path))
            ws = wb[_DATASET_SHEET] if _DATASET_SHEET in wb.sheetnames else wb.active
            updated = _apply_changes_to_sheet(ws, request, statement_type)
            total_updated += updated
            if updated > 0:
                wb.save(str(xlsx_path))
        except Exception as e:
            logger.warning("[apply-changes] Failed to update %s: %s", xlsx_path, e)

    return {"success": True, "rows_updated": total_updated}


def _apply_changes_to_sheet(ws, request: ApplyChangesRequest, statement_type: str) -> int:
    """Apply rename/add/delete operations to a single worksheet. Returns rows changed."""
    updated = 0

    # Build a map of label → row number for quick lookup
    # Column A is always the label column in our dataset format
    label_to_row: Dict[str, int] = {}
    for row_num in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=1)
        if cell.value and isinstance(cell.value, str):
            label_to_row[cell.value] = row_num

    # Renames
    for rename in request.renames:
        row_num = label_to_row.get(rename.old_label)
        if row_num is not None:
            ws.cell(row=row_num, column=1).value = rename.new_label
            label_to_row[rename.new_label] = row_num
            del label_to_row[rename.old_label]
            updated += 1

    # Deletions — collect row numbers descending so splicing doesn't shift indexes
    rows_to_delete = []
    for label in request.deletions:
        row_num = label_to_row.get(label)
        if row_num is not None:
            rows_to_delete.append(row_num)
    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)
        updated += 1

    # Additions — find the last data row and append with null values
    # Figure out how many period columns exist (all columns after column A)
    max_col = ws.max_column
    for label in request.additions:
        if label in label_to_row:
            continue  # already exists
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = label
        # Leave value columns as None (null) for all existing periods
        updated += 1

    return updated

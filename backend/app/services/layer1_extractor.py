"""
Python-side Excel extraction for the Layer 1 4-step pipeline.

Step A: extract_header_rows  — reads the first N rows of a sheet and returns
         them as a plain-text block for the AI column identifier.

Step C: extract_rows_with_metadata — reads the full sheet for the identified
         data column, returning one dict per row with value, formatting flags
         (bold, italic, indent), and the row label.

         rows_to_csv_with_metadata — serialises that list to a CSV string that
         the AI structured extractor can consume.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Font


# ── helpers ──────────────────────────────────────────────────────────────────

def _find_label_column(ws, start_row: int, end_row: Optional[int]) -> int:
    """
    Identify the most consistent label column by finding which column index
    most frequently holds the first non-empty text cell across the section.

    This prevents hidden account-code columns (which only have values on some
    rows) from winning over the real display-name column (which has values on
    every row).
    """
    from collections import Counter
    col_counts: Counter = Counter()
    max_row = end_row if end_row else start_row + 300

    for row in ws.iter_rows(min_row=start_row, max_row=max_row):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                col_counts[cell.column] += 1
                break

    return col_counts.most_common(1)[0][0] if col_counts else 1


def _effective_font(cell) -> Font:
    """Return the cell's font, falling back to a plain Font() if absent."""
    return cell.font if cell.font else Font()


def _indent_level(cell) -> int:
    """Return the cell's alignment indent level (0 if not set)."""
    if cell.alignment and cell.alignment.indent:
        return int(cell.alignment.indent)
    return 0


def _cell_value(cell) -> Any:
    return cell.value


# ── Step A ────────────────────────────────────────────────────────────────────

def extract_header_rows(
    filepath: str,
    sheet_name: str,
    n_rows: int = 150,
) -> str:
    """
    Open the workbook and return the first *n_rows* rows of *sheet_name* as a
    plain-text block with 1-based row numbers prepended.

    Format: "[row_num]\tcol1\tcol2\t..."

    Row numbers allow the AI column identifier (Step B) to return precise
    section_start_row / section_end_row values for multi-statement sheets.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    lines: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n_rows:
            break
        row_num = i + 1
        cells = [str(v) if v is not None else "" for v in row]
        lines.append(f"[{row_num}]\t" + "\t".join(cells))

    wb.close()
    return "\n".join(lines)


# ── Step C ────────────────────────────────────────────────────────────────────

def extract_rows_with_metadata(
    filepath: str,
    sheet_name: str,
    column_index: int,          # 1-based column index of the target period
    source_scaling: str,        # 'thousands' | 'millions' | 'actual_dollars'
    skip_rows: int = 0,
    section_start_row: int = 0,
    section_end_row: int = 0,
    label_col_override: Optional[int] = None,  # company-specific label column (skips _find_label_column)
) -> List[Dict[str, Any]]:
    """
    Read the sheet and return one dict per non-empty label row with:
      - label      : str   — text from the first non-empty cell in the row
      - label_col  : int   — 1-based column index where the label was found
      - value      : float | None
      - bold       : bool
      - italic     : bool
      - indent     : int   — alignment indent level (0 = leftmost)
      - row_index  : int   — 1-based sheet row number

    Row range: when section_start_row > 0, only rows in [section_start_row,
    section_end_row] are processed. This allows multi-statement sheets to be
    split correctly. Falls back to skip_rows when section_start_row is not set.

    Scaling is applied: values are normalised to actual dollars.
    Rows with no label text are skipped.
    """
    scale = _parse_scale(source_scaling)

    # Resolve effective row bounds
    if section_start_row > 0:
        start_row = section_start_row
    else:
        start_row = skip_rows + 1  # 1-based: skip_rows=3 means start at row 4

    end_row: Optional[int] = section_end_row if section_end_row > 0 else None

    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    ws = wb[sheet_name]

    label_column = label_col_override if label_col_override else _find_label_column(ws, start_row, end_row)

    rows: List[Dict[str, Any]] = []

    for row_num, row in enumerate(ws.iter_rows(), start=1):
        if row_num < start_row:
            continue
        if end_row is not None and row_num > end_row:
            break

        # Use the identified label column consistently for every row
        label_cell = ws.cell(row=row_num, column=label_column)
        label_val = label_cell.value
        label: Optional[str] = str(label_val).strip() if label_val is not None and str(label_val).strip() else None

        if label is None:
            continue  # blank row

        # Value from the target column
        value_cell = ws.cell(row=row_num, column=column_index)
        raw_val = _cell_value(value_cell)
        raw_str = str(raw_val).strip() if raw_val is not None else ""

        font = _effective_font(label_cell)
        is_bold = bool(font.bold)
        indent = _indent_level(label_cell)

        if raw_val is None or raw_str == "":
            # Genuinely empty cell — title/header row, skip it
            continue
        elif raw_str in ("-", "—", "–"):
            # Dash placeholder — could be a genuine zero OR a section header.
            # Skip non-bold, low-indent rows with a dash value: these are almost
            # always section headers that happen to have a dash in the value column.
            if not is_bold and indent <= 1:
                continue
            value: float = 0.0
        else:
            try:
                value = float(raw_str.replace(",", "").replace("(", "-").replace(")", "")) * scale
            except (ValueError, TypeError):
                # Non-numeric (e.g. "N/A", text) — skip
                continue

        rows.append({
            "label": label,
            "label_col": label_column,
            "value": value,
            "bold": is_bold,
            "italic": bool(font.italic),
            "indent": indent,
            "row_index": row_num,
        })

    wb.close()
    return rows


def rows_to_csv_with_metadata(rows: List[Dict[str, Any]]) -> str:
    """
    Serialise the output of extract_rows_with_metadata to a CSV string.

    Columns: row_index, label, value, bold, italic, indent
    """
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["row_index", "label", "value", "bold", "italic", "indent"],
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for r in rows:
        writer.writerow({
            "row_index": r["row_index"],
            "label": r["label"],
            "value": r["value"],
            "bold": r["bold"],
            "italic": r["italic"],
            "indent": r["indent"],
        })
    return output.getvalue()


# ── Display rows (full fidelity for template editor) ─────────────────────────

def extract_all_rows_for_display(
    filepath: str,
    sheet_name: str,
    column_index: int,
    source_scaling: str,
    skip_rows: int = 0,
    section_start_row: int = 0,
    section_end_row: int = 0,
    label_col_override: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Extract ALL rows between section boundaries for the template editor left panel.

    Unlike extract_rows_with_metadata, this includes blank rows and non-numeric
    rows. Every row in the section range is returned with:
      - row_index : int    — 1-based sheet row number
      - label     : str    — text from the identified label column, or "" for blank rows
      - value     : float | None — numeric value from target column (None if absent)
      - bold      : bool
      - italic    : bool
      - indent    : int    — alignment indent level
    """
    scale = _parse_scale(source_scaling)

    start_row = section_start_row if section_start_row > 0 else (skip_rows + 1)
    end_row: Optional[int] = section_end_row if section_end_row > 0 else None

    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    ws = wb[sheet_name]

    label_column = label_col_override if label_col_override else _find_label_column(ws, start_row, end_row)

    rows: List[Dict[str, Any]] = []

    for row_num, row in enumerate(ws.iter_rows(), start=1):
        if row_num < start_row:
            continue
        if end_row is not None and row_num > end_row:
            break

        # Use the consistent label column for every row
        label_cell = ws.cell(row=row_num, column=label_column)
        label_val = label_cell.value
        label = str(label_val).strip() if label_val is not None and str(label_val).strip() else ""

        # Value from the target column
        value_cell = ws.cell(row=row_num, column=column_index)
        raw_val = _cell_value(value_cell)
        value: Optional[float] = None
        if raw_val is not None:
            raw_str = str(raw_val).strip()
            if raw_str not in ("", "-", "—", "–"):
                try:
                    value = float(raw_str.replace(",", "").replace("(", "-").replace(")", "")) * scale
                except (ValueError, TypeError):
                    pass

        # Formatting from label cell (defaults for blank rows)
        if label_cell is not None:
            font = _effective_font(label_cell)
            is_bold = bool(font.bold)
            is_italic = bool(font.italic)
            indent = _indent_level(label_cell)
        else:
            is_bold = False
            is_italic = False
            indent = 0

        rows.append({
            "row_index": row_num,
            "label": label,
            "value": value,
            "bold": is_bold,
            "italic": is_italic,
            "indent": indent,
        })

    wb.close()

    # Trim trailing blank rows — keep all blanks in the middle, drop those at the end
    last_nonempty = len(rows) - 1
    while last_nonempty >= 0 and not rows[last_nonempty]["label"]:
        last_nonempty -= 1
    return rows[:last_nonempty + 1]


# ── Step B validation ─────────────────────────────────────────────────────────

def count_numeric_values_in_column(
    filepath: str,
    sheet_name: str,
    column_index: int,
    start_row: int = 1,
    end_row: Optional[int] = None,
) -> int:
    """
    Count rows that have a parseable numeric value in *column_index* (1-based).
    Used to validate the AI's column selection before running the full extraction.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    count = 0
    for row_num, row in enumerate(
        ws.iter_rows(min_col=column_index, max_col=column_index, values_only=True),
        start=1,
    ):
        if row_num < start_row:
            continue
        if end_row is not None and row_num > end_row:
            break
        raw = row[0]
        if raw is None:
            continue
        raw_str = str(raw).strip()
        if raw_str in ("", "-", "—", "–"):
            continue
        try:
            float(raw_str.replace(",", "").replace("(", "-").replace(")", ""))
            count += 1
        except (ValueError, TypeError):
            pass
    wb.close()
    return count


# ── internal ──────────────────────────────────────────────────────────────────

def _parse_scale(source_scaling: str) -> float:
    s = source_scaling.lower()
    if "thousand" in s:
        return 1_000.0
    if "million" in s:
        return 1_000_000.0
    return 1.0

"""
Excel processing service.
Handles reading Excel workbooks and converting sheets to CSVs for Layer 1 extraction.
PDF conversion has been removed — the browser now renders the original file client-side.
"""
import csv
import io
import re
from pathlib import Path
from typing import Dict, List

import openpyxl


def _safe_filename(name: str) -> str:
    """Sanitize a sheet name for use as a filename."""
    safe = re.sub(r'[\\/*?:"<>|]', "_", name)
    safe = safe.strip(". ")
    return safe or "sheet"


def get_sheet_names(filepath: str) -> List[str]:
    """
    Return an ordered list of visible sheet names from the workbook.

    Args:
        filepath: Absolute path to the .xlsx/.xls file.

    Returns:
        List of visible sheet names in workbook order.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = [
        name for name in wb.sheetnames
        if wb[name].sheet_state == "visible"
    ]
    wb.close()
    return names


def convert_to_csvs(filepath: str) -> Dict[str, str]:
    """
    Convert each visible sheet in an Excel workbook to CSV content using openpyxl.

    Args:
        filepath: Absolute path to the uploaded .xlsx / .xls file.

    Returns:
        Mapping of sheet_name → CSV content string.
        Content is NOT written to disk — the caller persists it.

    Notes:
        - Hidden and empty sheets are skipped.
        - All cell content is passed as-is; the AI prompt handles interpretation.
        - Merged cells: openpyxl returns the value in the top-left cell, None elsewhere.
    """
    source_path = Path(filepath)
    if not source_path.exists():
        raise FileNotFoundError(f"Upload file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    csv_contents: Dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.sheet_state != "visible":
            continue

        if ws.max_row is None or ws.max_row == 0:
            continue

        output = io.StringIO()
        writer = csv.writer(output)

        for row in ws.iter_rows(values_only=True):
            writer.writerow([("" if cell is None else str(cell)) for cell in row])

        csv_contents[sheet_name] = output.getvalue()

    wb.close()
    return csv_contents
